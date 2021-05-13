import pandas as pd
import openpyxl as openexcel
file = '/Users/omnisoportetecnico/Desktop/Omni/Dummy Folder/produceSales.xlsx'
excel_data = pd.ExcelFile(file)
print(excel_data.sheet_names)
# Exercise extraded by https://medium.com/analytics-vidhya/how-to-extract-information-from-your-excel-sheet-using-python-5f4f518aec49
#now we will verify if our xlsx file has been uploaded succesfully just showing the firt ten rows
df = excel_data.parse('Sheet1')
#print(df.info)
#print(df.head(10))

#time to work with the specific cells, for that pourpuse, we will use openpyxl for manipulate specific cells.
ps = openexcel.load_workbook('/Users/omnisoportetecnico/Desktop/Omni/Dummy Folder/produceSales.xlsx')
sheet = ps['Sheet1']
max_row_excel = sheet.max_row
print(sheet.max_row) 
#now, we will use openpyxl for return us the specific cell value
TotalInfo = {} # this will a nested dictionary https://www.programiz.com/python-programming/nested-dictionary

for i in range(2, max_row_excel + 1):
    # each row in the spreadsheet represents information for a particular purchase.
    produce = sheet['B' + str(i)].value
    cost_per_pound = sheet['C' + str(i)].value
    pounds_sold = sheet['D' + str(i)].value
    total_sales = sheet['E' + str(i)].value
    # Each row represents a fruit, so increment by the new corresponding values. 
    TotalInfo.setdefault(produce,{'Total_cost_per_pound': 0,
    'Total_pounds_sold': 0, 
    'Total_sales': 0,
    'Total_Purchase_Instances': 0})

    TotalInfo[produce]['Total_cost_per_pound'] += float(cost_per_pound)
    TotalInfo[produce]['Total_pounds_sold'] += int(pounds_sold)
    TotalInfo[produce]['Total_sales'] += int(total_sales)
    # Each row represents a fruit, so increment by one. 
    
    TotalInfo[produce]['Total_Purchase_Instances'] += 1
# the first column is B followed by C and so on.
# Each value in a cell is represented by a column letter and a row number. So #the first element in the sheet is B1, next column C1 and so on. This enables #to iterate over the entire cells.
print(TotalInfo['Apples'])

#dic = {produce,{'Total_cost_per_pound': 0,'Total_pounds_sold': 0, 'Total_sales': 0,'Total_Purchase_Instances': 0}}